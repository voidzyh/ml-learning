#!/usr/bin/env python3
'''
ML/DL 50周学习系统 - 核心控制器
功能: 读取Excel课表，追踪进度，实现 /today, /done, /status 等指令
'''

import json
import os
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Dict, Any, List

# 尝试导入 openpyxl，如果不可用则给出提示
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def _get_sr_manager():
    '''延迟导入间隔重复管理器，避免循环依赖'''
    try:
        import sys as _sys
        tools_path = str(Path(__file__).parent / 'tools')
        if tools_path not in _sys.path:
            _sys.path.insert(0, tools_path)
        from spaced_repetition import SpacedRepetitionManager
        return SpacedRepetitionManager()
    except ImportError:
        return None


class MLTutor:
    '''ML/DL学习系统核心类'''

    # 项目根目录
    BASE_DIR = Path(__file__).parent

    # 文件路径
    TRACKER_FILE = BASE_DIR / 'progress' / 'tracker.json'
    DAILY_SCHEDULE_FILE = BASE_DIR / 'data' / 'ML_DL_逐日课表_软工科班版.xlsx'
    Bilibili_RESOURCES_FILE = BASE_DIR / 'data' / 'B站ML_DL优质资源清单.xlsx'
    WEEKLY_OVERVIEW_FILE = BASE_DIR / 'data' / 'ML_DL_50周课表_软工科班版.xlsx'

    # Phase 映射
    PHASES = {
        0: '数学直觉 + NumPy/Pandas + sklearn入门',
        1: '经典ML（回归/分类/树/集成/聚类/PCA/特征工程）',
        2: 'DL基础（神经网络/PyTorch/优化/CNN/RNN/Attention）',
        3: 'Transformer + BERT/GPT + ViT + 推荐系统',
        4: 'LLM + RAG + 微调 + Agent + MLOps',
        5: '毕业项目 + 作品集 + 面试准备'
    }

    # 周名映射
    DAY_NAMES = ['周一', '周二', '周三', '周四', '周五', '周六']
    DAY_NAME_TO_INDEX = {name: i for i, name in enumerate(DAY_NAMES)}

    def __init__(self):
        self.tracker = self._load_tracker()
        self._schedule_cache = None
        self._bilibili_cache = None
        self._sr_manager = None  # 延迟初始化

    @property
    def sr_manager(self):
        '''延迟加载间隔重复管理器，首次访问时补建已完成天数的卡片'''
        if self._sr_manager is None:
            self._sr_manager = _get_sr_manager()
            if self._sr_manager:
                self._sr_manager.backfill_from_tracker(
                    self.tracker, self._load_schedule
                )
        return self._sr_manager

    def _load_tracker(self) -> dict:
        '''加载进度追踪文件'''
        if self.TRACKER_FILE.exists():
            with open(self.TRACKER_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return self._default_tracker()

    def _default_tracker(self) -> dict:
        '''返回默认的tracker结构'''
        return {
            'start_date': None,
            'current_week': 1,
            'current_day': 1,  # 1-6
            'streak': 0,
            'total_completed_days': 0,
            'total_skipped_days': 0,
            'phase': 0,
            'days': {},
            'projects': {
                'titanic-eda': {'status': 'not_started', 'github': '', 'week': 3},
                'numpy-lr': {'status': 'not_started', 'github': '', 'week': 4},
                'spam-classifier': {'status': 'not_started', 'github': '', 'week': 7},
                'customer-churn': {'status': 'not_started', 'github': '', 'week': 10},
                'kaggle-competition-1': {'status': 'not_started', 'github': '', 'week': 12},
                'numpy-neural-net': {'status': 'not_started', 'github': '', 'week': 13},
                'transfer-learning': {'status': 'not_started', 'github': '', 'week': 17},
                'imdb-sentiment': {'status': 'not_started', 'github': '', 'week': 18},
                'mnist-cnn-99': {'status': 'not_started', 'github': '', 'week': 20},
                'minigpt': {'status': 'not_started', 'github': '', 'week': 22},
                'bert-classification': {'status': 'not_started', 'github': '', 'week': 24},
                'recommendation-web': {'status': 'not_started', 'github': '', 'week': 29},
                'rag-qa-system': {'status': 'not_started', 'github': '', 'week': 35},
                'llm-lora-finetune': {'status': 'not_started', 'github': '', 'week': 36},
                'mlops-pipeline': {'status': 'not_started', 'github': '', 'week': 41},
                'capstone-project': {'status': 'not_started', 'github': '', 'week': 45},
                'kaggle-featured': {'status': 'not_started', 'github': '', 'week': 47}
            },
            'blogs': [],
            'quiz_scores': [],
            'knowledge_gaps': []
        }

    def _save_tracker(self):
        '''保存进度追踪文件'''
        self.TRACKER_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(self.TRACKER_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.tracker, f, ensure_ascii=False, indent=2)

    def _get_day_key(self, week: int, day: int) -> str:
        '''生成天的键名，如 W1D1'''
        return f'W{week}D{day}'

    def _load_schedule(self) -> List[Dict]:
        '''加载逐日课表Excel'''
        if not EXCEL_AVAILABLE:
            print('⚠️  需要安装 openpyxl: pip install openpyxl')
            return []

        if not self.DAILY_SCHEDULE_FILE.exists():
            print(f'⚠️  找不到课表文件: {self.DAILY_SCHEDULE_FILE}')
            print('   请将 ML_DL_逐日课表_软工科班版.xlsx 放入 data/ 目录')
            return []

        if self._schedule_cache is not None:
            return self._schedule_cache

        schedule = []
        wb = load_workbook(self.DAILY_SCHEDULE_FILE, data_only=True)
        ws = wb.active

        current_phase = 0
        current_week = None

        for row in ws.iter_rows(min_row=2, values_only=False):
            if not row[0]:  # 空行跳过
                continue

            # 检查是否是Phase标题行（合并单元格）
            if row[0].value and isinstance(row[0].value, str) and row[0].value.startswith('Phase'):
                # 解析 Phase 号
                try:
                    phase_num = int(row[0].value.split()[1].rstrip('：'))
                    current_phase = phase_num
                except (IndexError, ValueError):
                    pass
                continue

            # 获取周数
            week_val = row[0].value
            if week_val is None or week_val == '':
                continue

            try:
                week_num = int(week_val)
            except (ValueError, TypeError):
                # 可能是标题行或其他非数据行
                continue

            if week_num > 50:  # 超出范围
                continue

            # 获取天
            day_val = row[1].value
            if not day_val or day_val not in self.DAY_NAME_TO_INDEX:
                continue

            day_index = self.DAY_NAME_TO_INDEX[day_val] + 1  # 转换为1-6

            # 解析数据行
            schedule_item = {
                'phase': current_phase,
                'week': week_num,
                'day': day_index,
                'day_name': day_val,
                'morning_theory': self._get_cell_value(row, 3),
                'afternoon_practice': self._get_cell_value(row, 4),
                'deliverables': self._get_cell_value(row, 5),
                'cumulative_hours': self._get_cell_value(row, 6)
            }
            schedule.append(schedule_item)

        self._schedule_cache = schedule
        return schedule

    def _get_cell_value(self, row, index: int) -> str:
        '''安全获取单元格值'''
        try:
            if index < len(row) and row[index].value:
                return str(row[index].value).strip()
        except (IndexError, AttributeError):
            pass
        return ''

    def _get_bilibili_resources(self, week: int) -> List[Dict]:
        '''获取指定周的B站资源'''
        if not EXCEL_AVAILABLE:
            return []

        if not self.Bilibili_RESOURCES_FILE.exists():
            return []

        resources = []
        wb = load_workbook(self.Bilibili_RESOURCES_FILE, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=False):
            if not row[0]:  # 空行
                continue

            # 解析'对应课表周'列
            week_range = self._get_cell_value(row, 6)
            if self._week_in_range(week, week_range):
                resources.append({
                    'name': self._get_cell_value(row, 2),
                    'uploader': self._get_cell_value(row, 3),
                    'keyword': self._get_cell_value(row, 4),
                    'description': self._get_cell_value(row, 5),
                    'priority': self._get_cell_value(row, 7),
                    'duration': self._get_cell_value(row, 8)
                })

        return resources

    def _week_in_range(self, week: int, range_str: str) -> bool:
        '''判断周是否在范围字符串内'''
        if not range_str:
            return False

        range_str = range_str.replace('第', '').replace('周', '').strip()

        # 处理多种格式
        if '-' in range_str:
            try:
                parts = range_str.split('-')
                # 处理 '1-32' 或 '1-32周'
                start = int(''.join(filter(str.isdigit, parts[0])))
                end = int(''.join(filter(str.isdigit, parts[1])))
                return start <= week <= end
            except ValueError:
                pass

        # 处理单个数字
        try:
            return int(range_str.split()[0]) == week
        except (ValueError, IndexError):
            pass

        # 处理 'P0-P3' 格式
        if 'P' in range_str:
            phase_ranges = {
                'P0': (1, 3), 'P1': (4, 12), 'P2': (13, 20),
                'P3': (21, 32), 'P4': (33, 42), 'P5': (43, 50)
            }
            for phase_key, (start_w, end_w) in phase_ranges.items():
                if phase_key in range_str:
                    return start_w <= week <= end_w

        return False

    # ========== 公共方法 ==========

    def ensure_daily_note(self, plan: Dict = None) -> Optional[str]:
        '''如果今日 Obsidian 日记不存在，自动创建并返回路径'''
        today = datetime.now().strftime('%Y-%m-%d')
        daily_dir = Path(__file__).parent / 'obsidian-vault' / '00-Daily'
        daily_file = daily_dir / f'{today}.md'

        if daily_file.exists():
            return str(daily_file)

        if plan is None:
            plan = self.get_today_plan()

        daily_dir.mkdir(parents=True, exist_ok=True)
        item = plan.get('schedule_item') or {}
        week = plan['week']
        day_name = plan['day_name']
        phase = plan['phase']
        phase_name = plan['phase_name']

        # 生成今日重点
        focus_points = self._get_focus_points(item) if item else []
        focus_section = '\n'.join(f'- {p}' for p in focus_points) if focus_points else '- （暂无）'

        # 生成完成情况 checkbox
        deliverables = item.get('deliverables', '') if item else ''
        checklist = '- [ ] 上午理论\n- [ ] 下午实践'
        if deliverables:
            checklist += f'\n- [ ] 今日交付: {deliverables}'

        # 生成明日预告
        tomorrow_preview = self._get_tomorrow_preview()

        content = f"""---
date: {today}
week: {week}
day: {plan['day']}
phase: {phase}
tags: [daily/week-{week}, phase-{phase}]
---

# {today} · 第{week}周·{day_name} | Phase {phase} {phase_name}

## 今日重点
{focus_section}

## 完成情况
{checklist}

## 上午·理论
{item.get('morning_theory', '暂无内容') if item else '暂无内容'}

### 笔记


## 下午·实践
{item.get('afternoon_practice', '暂无内容') if item else '暂无内容'}

### 笔记


## 今日交付
{deliverables}

## 收获与疑问

### 今日收获


### 遗留疑问


## 明日预告
> {tomorrow_preview}
"""
        daily_file.write_text(content, encoding='utf-8')
        return str(daily_file)

    def update_daily_note_on_done(self, done_result: Dict) -> Optional[str]:
        '''打卡后更新今日日记，追加打卡总结区块

        Args:
            done_result: mark_done() 的返回值

        Returns:
            日记文件路径，如果文件不存在或已包含总结则返回 None
        '''
        today = datetime.now().strftime('%Y-%m-%d')
        daily_file = Path(__file__).parent / 'obsidian-vault' / '00-Daily' / f'{today}.md'

        if not daily_file.exists():
            return None

        content = daily_file.read_text(encoding='utf-8')

        # 幂等性：已有打卡总结则跳过
        if '## 打卡总结' in content:
            return None

        now = datetime.now()
        week = done_result['week']
        day = done_result['day']
        streak = done_result['streak']
        progress = done_result['progress']
        new_cards = done_result.get('new_review_cards', [])

        # 构建打卡总结区块
        summary_lines = [
            '',
            '## 打卡总结',
            f'> 完成时间: {now.strftime("%H:%M")} | '
            f'连续学习: {streak}天 | '
            f'总进度: {progress:.1f}%',
            '',
        ]

        # 新建复习卡片
        if new_cards:
            summary_lines.append('### 今日新建复习卡片')
            for card in new_cards:
                summary_lines.append(f'- [[{card}]] (明天复习)')
            summary_lines.append('')

        # 从课表提取今日核心概念
        concepts = []
        for item in self._load_schedule():
            if item['week'] == week and item['day'] == day:
                morning = item.get('morning_theory', '')
                if morning:
                    for part in morning.split('•'):
                        c = part.strip()
                        if c and len(c) > 2 and len(c) < 50:
                            concepts.append(c)
                break

        if concepts:
            summary_lines.append('### 今日核心概念')
            for c in concepts:
                summary_lines.append(f'- [[{c}]]')
            summary_lines.append('')

        summary_block = '\n'.join(summary_lines) + '\n'

        # 自动勾选完成情况 checkbox
        content = content.replace('- [ ] 上午理论', '- [x] 上午理论')
        content = content.replace('- [ ] 下午实践', '- [x] 下午实践')
        # 勾选今日交付（如果有）
        content = re.sub(r'- \[ \] 今日交付:', '- [x] 今日交付:', content)

        # 在 "## 收获与疑问" 之前插入打卡总结
        if '## 收获与疑问' in content:
            content = content.replace(
                '## 收获与疑问',
                summary_block + '## 收获与疑问'
            )
        else:
            # 如果没有"收获与疑问"区块，追加到末尾
            content = content.rstrip() + '\n' + summary_block

        daily_file.write_text(content, encoding='utf-8')
        return str(daily_file)

    def get_today_plan(self) -> Dict[str, Any]:
        '''获取今日学习计划 (/today 指令)'''
        week = self.tracker['current_week']
        day = self.tracker['current_day']
        phase = self.tracker['phase']

        schedule = self._load_schedule()
        today_item = None

        for item in schedule:
            if item['week'] == week and item['day'] == day:
                today_item = item
                break

        resources = self._get_bilibili_resources(week)

        # 获取今日到期的复习卡片
        due_reviews = []
        if self.sr_manager:
            due_reviews = self.sr_manager.get_due_cards()

        return {
            'week': week,
            'day': day,
            'day_name': self.DAY_NAMES[day - 1],
            'phase': phase,
            'phase_name': self.PHASES.get(phase, ''),
            'schedule_item': today_item,
            'bilibili_resources': resources,
            'due_reviews': due_reviews
        }

    def mark_done(self) -> Dict[str, Any]:
        '''标记今日完成 (/done 指令)'''
        week = self.tracker['current_week']
        day = self.tracker['current_day']
        day_key = self._get_day_key(week, day)

        # 幂等性检查：防止重复标记
        if day_key in self.tracker['days']:
            existing = self.tracker['days'][day_key]
            if existing.get('status') == 'done':
                return {
                    'error': f'第{week}周第{day}天已经完成过了',
                    'completed_at': existing.get('completed_at'),
                    'week': week,
                    'day': day,
                    'progress': (self.tracker['total_completed_days'] / 300) * 100,
                    'streak': self.tracker['streak'],
                    'next_week': self.tracker['current_week'],
                    'next_day': self.tracker['current_day'],
                    'is_saturday': False,
                    'new_review_cards': []
                }

        # 更新当天状态
        self.tracker['days'][day_key] = {
            'status': 'done',
            'completed_at': datetime.now().isoformat(),
            'notes': ''
        }

        self.tracker['total_completed_days'] += 1
        self.tracker['streak'] += 1

        # 计算进度百分比
        total_days = 50 * 6  # 50周 × 6天
        progress = (self.tracker['total_completed_days'] / total_days) * 100

        # 自动创建间隔重复卡片
        new_cards = []
        if self.sr_manager:
            for item in self._load_schedule():
                if item['week'] == week and item['day'] == day:
                    mt = item.get('morning_theory', '')
                    if mt:
                        new_cards = self.sr_manager.create_cards_from_day(week, day, mt)
                    break

        # 推进到下一天
        is_saturday = (day == 6)
        if is_saturday:
            # 周六完成，推进到下一周周一
            next_week = self.tracker['current_week'] + 1

            # 检查是否超出50周
            if next_week > 50:
                # 已完成全部50周，不再推进
                self._save_tracker()
                return {
                    'week': week,
                    'day': day,
                    'progress': 100.0,
                    'streak': self.tracker['streak'],
                    'next_week': 50,
                    'next_day': 6,
                    'is_saturday': True,
                    'new_review_cards': new_cards,
                    'weekly_review_generated': True,
                    'completion_rate': 1.0,
                    'course_completed': True
                }

            self.tracker['current_week'] = next_week
            self.tracker['current_day'] = 1
            # 更新Phase
            self._update_phase()

            # 检查本周完成率，决定是否自动生成周回顾
            week_overview = self.get_week_overview(week)
            completion_rate = week_overview['completed'] / 6

            if completion_rate >= 0.5:  # 至少完成一半
                review_data = self.generate_weekly_review(week)
                self.save_weekly_review(review_data)
                weekly_review_generated = True
            else:
                weekly_review_generated = False
        else:
            self.tracker['current_day'] += 1
            weekly_review_generated = False

        self._save_tracker()

        return {
            'week': week,
            'day': day,
            'progress': progress,
            'streak': self.tracker['streak'],
            'next_week': self.tracker['current_week'],
            'next_day': self.tracker['current_day'],
            'is_saturday': is_saturday,
            'new_review_cards': new_cards,
            'weekly_review_generated': weekly_review_generated,
            'completion_rate': completion_rate if is_saturday else None
        }

    def mark_skip(self, reason: str = '') -> Dict[str, Any]:
        '''跳过今天 (/skip 指令)'''
        week = self.tracker['current_week']
        day = self.tracker['current_day']
        day_key = self._get_day_key(week, day)

        self.tracker['days'][day_key] = {
            'status': 'skipped',
            'reason': reason,
            'reschedule': True
        }

        self.tracker['total_skipped_days'] += 1
        self.tracker['streak'] = 0

        # 不推进进度指针
        self._save_tracker()

        return {'week': week, 'day': day, 'reason': reason}

    def get_status(self) -> Dict[str, Any]:
        '''获取总进度 (/status 指令)'''
        total_days = 50 * 6
        progress = (self.tracker['total_completed_days'] / total_days) * 100

        # 统计项目完成情况
        completed_projects = sum(
            1 for p in self.tracker['projects'].values()
            if p.get('status') == 'done'
        )

        # 统计待补天数
        pending_makeup = sum(
            1 for d in self.tracker['days'].values()
            if d.get('status') == 'skipped' and d.get('reschedule')
        )

        return {
            'current_week': self.tracker['current_week'],
            'current_day': self.tracker['current_day'],
            'day_name': self.DAY_NAMES[self.tracker['current_day'] - 1],
            'phase': self.tracker['phase'],
            'phase_name': self.PHASES.get(self.tracker['phase'], ''),
            'progress': progress,
            'streak': self.tracker['streak'],
            'total_completed': self.tracker['total_completed_days'],
            'total_skipped': self.tracker['total_skipped_days'],
            'completed_projects': completed_projects,
            'total_projects': len(self.tracker['projects']),
            'pending_makeup': pending_makeup,
            'blogs': len(self.tracker.get('blogs', [])),
            'quiz_count': len(self.tracker.get('quiz_scores', [])),
            'start_date': self.tracker.get('start_date')
        }

    def get_week_overview(self, week: Optional[int] = None) -> Dict[str, Any]:
        '''获取本周概览 (/week 指令)'''
        if week is None:
            week = self.tracker['current_week']

        # 验证周数范围
        if week < 1 or week > 50:
            return {
                'error': f'周数超出范围（1-50）: {week}',
                'week': week,
                'days': []
            }

        schedule = self._load_schedule()
        week_days = []

        for day_idx in range(1, 7):  # 1-6
            day_key = self._get_day_key(week, day_idx)
            day_status = self.tracker['days'].get(day_key, {}).get('status', 'pending')

            # 查找课表内容
            schedule_item = None
            for item in schedule:
                if item['week'] == week and item['day'] == day_idx:
                    schedule_item = item
                    break

            week_days.append({
                'day': day_idx,
                'day_name': self.DAY_NAMES[day_idx - 1],
                'status': day_status,
                'schedule': schedule_item
            })

        return {
            'week': week,
            'days': week_days
        }

    def set_start_date(self, date_str: str):
        '''设置开始日期'''
        self.tracker['start_date'] = date_str
        self._save_tracker()

    def jump_to(self, week: int, day: int):
        '''跳转到指定周和天（用于追赶进度）'''
        # 验证周数范围
        if week < 1 or week > 50:
            raise ValueError(f'周数必须在 1-50 之间，当前值: {week}')
        # 验证天数范围
        if day < 1 or day > 6:
            raise ValueError(f'天数必须在 1-6 之间，当前值: {day}')

        self.tracker['current_week'] = week
        self.tracker['current_day'] = day
        self._update_phase()
        self._save_tracker()

    def _get_tomorrow_preview(self) -> str:
        '''获取明日学习内容预告'''
        week = self.tracker['current_week']
        day = self.tracker['current_day']

        # 计算下一天
        if day < 6:
            next_week, next_day = week, day + 1
        else:
            next_week, next_day = week + 1, 1

        if next_week > 50:
            return '已完成全部50周学习计划！'

        schedule = self._load_schedule()
        for item in schedule:
            if item['week'] == next_week and item['day'] == next_day:
                theory = item.get('morning_theory', '')
                # 截取摘要（取第一个要点）
                summary = theory.split('•')[0].strip() if theory else ''
                if len(summary) > 60:
                    summary = summary[:60] + '...'
                day_name = self.DAY_NAMES[next_day - 1]
                return f'第{next_week}周·{day_name}: {summary}'

        return f'第{next_week}周·{self.DAY_NAMES[next_day - 1]}'

    def _get_focus_points(self, schedule_item: Dict) -> list:
        '''从课表项中提取今日重点列表'''
        points = []
        morning = schedule_item.get('morning_theory', '')
        afternoon = schedule_item.get('afternoon_practice', '')

        if morning:
            # 按 • 分隔提取理论要点
            parts = morning.split('•')
            for part in parts:
                concept = part.strip()
                if concept and len(concept) > 2:
                    # 去掉过长的描述，只取核心
                    if len(concept) > 50:
                        concept = concept[:50] + '...'
                    points.append(f'理解 **{concept}** 的核心思想')

        if afternoon:
            # 取实践摘要（只取第一行）
            first_line = afternoon.split('\n')[0].strip()
            summary = first_line.split('•')[0].strip()
            if summary and len(summary) > 2:
                if len(summary) > 50:
                    summary = summary[:50] + '...'
                points.append(f'实践: {summary}')

        return points

    def _update_phase(self):
        '''根据当前周更新Phase'''
        week = self.tracker['current_week']
        if week <= 3:
            self.tracker['phase'] = 0
        elif week <= 12:
            self.tracker['phase'] = 1
        elif week <= 20:
            self.tracker['phase'] = 2
        elif week <= 32:
            self.tracker['phase'] = 3
        elif week <= 42:
            self.tracker['phase'] = 4
        else:
            self.tracker['phase'] = 5

    def generate_quiz(self, topic: str = None, count: int = 5) -> Dict[str, Any]:
        '''生成测验题 (/quiz 指令)

        Args:
            topic: 主题（如 'linear-algebra', 'gradient-descent'），None则自动推断
            count: 题目数量，默认5道
        '''
        # 如果没有指定主题，根据当前周和Phase推断
        if topic is None:
            topic = self._infer_topic_from_progress()

        questions = self._get_questions_for_topic(topic, count)

        return {
            'topic': topic,
            'count': len(questions),
            'questions': questions
        }

    def _infer_topic_from_progress(self) -> str:
        '''根据当前进度推断测验主题'''
        week = self.tracker['current_week']
        phase = self.tracker['phase']

        # Phase 0: 数学基础
        if phase == 0:
            if week == 1:
                return 'linear-algebra'
            elif week == 2:
                return 'calculus'
            else:
                return 'numpy-basics'

        # Phase 1: 经典ML
        elif phase == 1:
            week_topics = {
                4: 'linear-regression',
                5: 'logistic-regression',
                6: 'decision-tree',
                7: 'naive-bayes',
                8: 'svm',
                9: 'ensemble',
                10: 'classification-metrics',
                11: 'clustering',
                12: 'pca'
            }
            return week_topics.get(week, 'ml-basics')

        # Phase 2: DL基础
        elif phase == 2:
            if week <= 14:
                return 'neural-networks'
            elif week <= 17:
                return 'optimization'
            elif week <= 20:
                return 'cnn'

        # Phase 3: Transformer
        elif phase == 3:
            if week <= 24:
                return 'attention'
            elif week <= 28:
                return 'transformer'
            else:
                return 'recommendation'

        # Phase 4: LLM
        elif phase == 4:
            if week <= 36:
                return 'bert'
            elif week <= 38:
                return 'gpt'
            else:
                return 'rag'

        return 'ml-basics'

    def _get_questions_for_topic(self, topic: str, count: int) -> List[Dict]:
        '''获取指定主题的题目'''
        all_questions = QUIZ_BANK.get(topic, QUIZ_BANK['ml-basics'])
        # 随机选择题目，但不超过可用数量
        import random
        return random.sample(all_questions, min(count, len(all_questions)))

    def save_quiz_score(self, topic: str, score: float, total: int):
        '''保存测验成绩'''
        self.tracker['quiz_scores'].append({
            'topic': topic,
            'score': score,
            'total': total,
            'percentage': (score / total) * 100 if total > 0 else 0,
            'date': datetime.now().isoformat()
        })
        self._save_tracker()

    def generate_weekly_review(self, week: int = None) -> Dict[str, Any]:
        '''生成周回顾 (/review 指令)

        Args:
            week: 周数，None则使用当前周
        '''
        if week is None:
            week = self.tracker['current_week']

        # 获取本周6天的状态
        overview = self.get_week_overview(week)

        # 统计完成情况
        completed = sum(1 for d in overview['days'] if d['status'] == 'done')
        skipped = sum(1 for d in overview['days'] if d['status'] == 'skipped')
        pending = sum(1 for d in overview['days'] if d['status'] == 'pending')

        # 提取本周学习的核心概念
        concepts = self._extract_week_concepts(week)

        # 生成自测题
        quiz_topic = self._infer_topic_from_week(week)
        quiz_questions = self._get_questions_for_topic(quiz_topic, 5)

        # 识别薄弱点
        weak_points = self._identify_weak_points(week, overview)

        return {
            'week': week,
            'completed': completed,
            'skipped': skipped,
            'pending': pending,
            'completion_rate': (completed / 6) * 100,
            'concepts': concepts,
            'quiz_questions': quiz_questions,
            'weak_points': weak_points
        }

    def _infer_topic_from_week(self, week: int) -> str:
        '''根据周数推断主题'''
        if week <= 3:
            return 'linear-algebra' if week == 1 else 'calculus' if week == 2 else 'numpy-basics'
        elif week <= 12:
            week_topics = {
                4: 'linear-regression', 5: 'logistic-regression', 6: 'decision-tree',
                7: 'naive-bayes', 8: 'svm', 9: 'ensemble', 10: 'classification-metrics',
                11: 'clustering', 12: 'pca'
            }
            return week_topics.get(week, 'ml-basics')
        elif week <= 20:
            return 'neural-networks' if week <= 17 else 'cnn'
        elif week <= 32:
            return 'attention' if week <= 24 else 'transformer'
        else:
            return 'bert'

    def _extract_week_concepts(self, week: int) -> List[str]:
        '''从本周课表中提取核心概念'''
        schedule = self._load_schedule()
        week_items = [item for item in schedule if item['week'] == week]

        concepts = []
        for item in week_items:
            # 从上午理论中提取关键概念
            theory = item.get('morning_theory', '')
            if theory:
                # 简单提取：按分隔符分割
                parts = theory.split('•')
                for part in parts[:2]:  # 每天最多取2个
                    concept = part.strip()
                    if concept and len(concept) > 2 and len(concept) < 50:
                        concepts.append(f'{item['day_name']}: {concept}')

        return concepts[:10]  # 最多返回10个

    def _identify_weak_points(self, week: int, overview: Dict) -> List[str]:
        '''识别本周薄弱点'''
        weak_points = []

        # 检查跳过的天数
        for day_info in overview['days']:
            if day_info['status'] == 'skipped':
                weak_points.append(f'{day_info['day_name']} 内容待补')

        # 检查未完成的天数
        for day_info in overview['days']:
            if day_info['status'] == 'pending' and day_info.get('schedule'):
                theory = day_info['schedule'].get('morning_theory', '')
                if theory:
                    weak_points.append(f'{day_info['day_name']}: {theory[:30]}...')

        return weak_points

    def save_weekly_review(self, review_data: Dict):
        '''保存周回顾到文件'''
        reviews_dir = self.BASE_DIR / 'obsidian-vault' / '04-Reviews'
        reviews_dir.mkdir(parents=True, exist_ok=True)

        filename = f'Week-{review_data['week']:02d}-Review.md'
        filepath = reviews_dir / filename

        # 生成Markdown内容
        content = self._format_review_as_markdown(review_data)

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(content)

        return str(filepath)

    def _format_review_as_markdown(self, review: Dict) -> str:
        '''将周回顾格式化为Markdown'''
        lines = [
            f'# 第{review['week']}周回顾',
            '',
            f'> 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}',
            '',
            '## 📊 完成情况',
            '',
            f'- ✅ 已完成: {review['completed']}/6 天',
            f'- ⏭️  跳过: {review['skipped']} 天',
            f'- ⬜ 待完成: {review['pending']} 天',
            f'- 📈 完成率: {review['completion_rate']:.1f}%',
            ''
        ]

        # 核心概念
        if review['concepts']:
            lines.extend([
                '## 📚 本周核心概念',
                ''
            ])
            for concept in review['concepts']:
                lines.append(f'- {concept}')
            lines.append('')

        # 自测题
        if review['quiz_questions']:
            lines.extend([
                '## 📝 自测题',
                ''
            ])
            for i, q in enumerate(review['quiz_questions'], 1):
                lines.append(f'### {i}. {q['question']}')
                lines.append(f'**类型**: {q['type']}')
                if q.get('options'):
                    for opt in q['options']:
                        lines.append(f'- {opt}')
                lines.append(f'**答案**: {q['answer']}')
                lines.append('')
            lines.append('')

        # 薄弱点
        if review['weak_points']:
            lines.extend([
                '## ⚠️ 待补强内容',
                ''
            ])
            for point in review['weak_points']:
                lines.append(f'- [ ] {point}')
            lines.append('')

        lines.extend([
            '---',
            '',
            '*建议: 完成自测题后，将薄弱点更新到 `knowledge-gaps.md`*'
        ])

        return '\n'.join(lines)


# ========== 题库 ==========

QUIZ_BANK = {
    'ml-basics': [
        {
            'type': '概念题',
            'question': '什么是机器学习？用一句话解释。',
            'answer': '机器学习是通过数据训练模型，让计算机从经验中自动改进性能的算法，而不需要显式编程。'
        },
        {
            'type': '概念题',
            'question': '监督学习和无监督学习的核心区别是什么？',
            'answer': '监督学习使用有标签的数据进行训练（输入+正确答案），无监督学习使用无标签数据，让算法自己发现数据中的模式。'
        },
        {
            'type': '对比题',
            'question': '训练集、验证集、测试集的区别和用途？',
            'answer': '训练集用于模型训练；验证集用于调参和模型选择，防止过拟合；测试集用于最终评估模型泛化能力，在模型开发过程中不参与任何决策。'
        },
        {
            'type': '场景题',
            'question': '如果你的模型在训练集上表现很好，但在测试集上表现很差，可能是什么问题？如何解决？',
            'answer': '这是过拟合问题。解决方法：1) 增加训练数据；2) 使用正则化；3) 简化模型；4) 使用交叉验证；5) 早停。'
        },
        {
            'type': '代码题',
            'question': '用 sklearn 将数据分割为训练集和测试集的代码是什么？',
            'answer': '```python\nfrom sklearn.model_selection import train_test_split\nX_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)\n```'
        }
    ],

    'linear-algebra': [
        {
            'type': '概念题',
            'question': '用编程思维理解：什么是向量？',
            'answer': '向量可以理解为一个一维数组，包含多个数值。在几何上，它是从原点指向空间某点的有向线段；在编程中，它是 List 或 NumPy Array。'
        },
        {
            'type': '概念题',
            'question': '矩阵乘法的几何意义是什么？',
            'answer': '矩阵乘法代表线性变换——对空间进行旋转、缩放、剪切等操作，把一个向量映射到另一个位置。'
        },
        {
            'type': '对比题',
            'question': '点积 和叉积 的区别？',
            'answer': '点积结果是一个标量，衡量两个向量的相似度/投影长度；叉积结果是一个向量，垂直于原来两个向量构成的平面，表示面积和方向。'
        },
        {
            'type': '代码题',
            'question': '用 NumPy 实现矩阵乘法',
            'answer': '```python\nimport numpy as np\nA = np.array([[1, 2], [3, 4]])\nB = np.array([[5, 6], [7, 8]])\nC = A @ B  # 或 np.dot(A, B) 或 np.matmul(A, B)\n```'
        },
        {
            'type': '场景题',
            'question': '在机器学习中，为什么我们需要矩阵求逆？',
            'answer': '在正规方程（Normal Equation）求解线性回归时，需要求逆来直接计算最优参数：θ = (X^T X)^(-1) X^T y。但实际常用梯度下降，因为求逆计算成本高且可能不可逆。'
        }
    ],

    'calculus': [
        {
            'type': '概念题',
            'question': '导数和梯度的区别？',
            'answer': '导数是标量函数的变化率，描述一维函数在某点的斜率；梯度是向量函数的导数，指向函数增长最快的方向，其大小是增长率。'
        },
        {
            'type': '概念题',
            'question': '为什么梯度下降能找到最小值？',
            'answer': '因为梯度指向函数增长最快的方向，所以负梯度方向就是下降最快的方向。沿着负梯度移动，函数值会减小，逐步逼近局部最小值。'
        },
        {
            'type': '对比题',
            'question': '学习率太大或太小会有什么问题？',
            'answer': '学习率太大：可能无法收敛，在最小值附近震荡甚至发散；学习率太小：收敛速度极慢，训练时间过长，可能陷入局部最优。'
        },
        {
            'type': '代码题',
            'question': '用 NumPy 实现简单的一维梯度下降',
            'answer': '```python\nimport numpy as np\n\ndef f(x): return x**2  # 目标函数\ndef df(x): return 2*x  # 导数\n\nx = 10.0  # 初始值\nlr = 0.1  # 学习率\nfor _ in range(100):\n    x = x - lr * df(x)\nprint(x)  # 输出接近0\n```'
        },
        {
            'type': '场景题',
            'question': '什么是局部最小值和全局最小值？如何避免陷入局部最优？',
            'answer': '局部最小值是在某个邻域内最小的点；全局最小值是整个定义域内最小的点。避免方法：随机初始化、使用动量、Adam优化器、模拟退火等。'
        }
    ],

    'numpy-basics': [
        {
            'type': '代码题',
            'question': '创建一个 3x3 的全零矩阵',
            'answer': '```python\nimport numpy as np\nnp.zeros((3, 3))\n```'
        },
        {
            'type': '代码题',
            'question': 'NumPy 数组的切片：获取第2-3行，第1-2列',
            'answer': '```python\narr[1:3, 0:2]  # Python索引从0开始\n```'
        },
        {
            'type': '对比题',
            'question': 'reshape() 和 resize() 的区别？',
            'answer': 'reshape 返回新数组，原数组不变；resize 直接修改原数组（in-place）或返回新数组（视调用方式）。'
        },
        {
            'type': '概念题',
            'question': '什么是广播（Broadcasting）？',
            'answer': '广播是 NumPy 对不同形状数组进行算术运算的机制。较小的数组会自动扩展以匹配较大数组的形状，无需显式复制数据。'
        },
        {
            'type': '代码题',
            'question': '计算两个数组的欧氏距离',
            'answer': '```python\nimport numpy as np\na = np.array([1, 2, 3])\nb = np.array([4, 5, 6])\ndist = np.linalg.norm(a - b)  # 或 np.sqrt(((a-b)**2).sum())\n```'
        }
    ],

    'linear-regression': [
        {
            'type': '概念题',
            'question': '线性回归的损失函数是什么？为什么用平方误差？',
            'answer': 'MSE = (1/n)Σ(y_i - ŷ_i)²。平方误差的优点：1) 可导；2) 惩罚大误差更重；3) 在高斯噪声假设下是最大似然估计。'
        },
        {
            'type': '概念题',
            'question': 'R² 是什么？它的取值范围和含义？',
            'answer': 'R²（决定系数）衡量模型解释方差的比例，范围通常在0到1之间。R²=1表示完美拟合，R²=0表示模型和简单取均值一样，R²<0表示模型比均值还差。'
        },
        {
            'type': '对比题',
            'question': '简单线性回归和多元线性回归的区别？',
            'answer': '简单线性回归只有一个特征变量；多元线性回归有多个特征变量，需要考虑特征之间的相关性和多重共线性问题。'
        },
        {
            'type': '代码题',
            'question': '用 sklearn 训练一个线性回归模型',
            'answer': '```python\nfrom sklearn.linear_model import LinearRegression\nmodel = LinearRegression()\nmodel.fit(X_train, y_train)\ny_pred = model.predict(X_test)\n```'
        },
        {
            'type': '场景题',
            'question': '线性回归的假设条件有哪些？如何检验？',
            'answer': '假设：1) 线性关系；2) 误差独立同分布；3) 误差同方差；4) 无多重共线性；5) 误差正态分布。检验方法：残差图、Q-Q图、VIF、Durbin-Watson检验等。'
        }
    ],

    'logistic-regression': [
        {
            'type': '概念题',
            'question': 'Sigmoid 函数的作用是什么？',
            'answer': 'Sigmoid 将任意实数映射到 (0,1) 区间，输出可解释为概率。公式：σ(z) = 1/(1+e^(-z))'
        },
        {
            'type': '对比题',
            'question': '线性回归和逻辑回归的区别？',
            'answer': '线性回归预测连续值，使用MSE损失；逻辑回归预测概率（二分类），使用交叉熵损失，输出经过Sigmoid激活。'
        },
        {
            'type': '概念题',
            'question': '交叉熵损失的含义是什么？',
            'answer': '衡量预测概率分布与真实标签分布之间的差异。对于二分类：L = -[y·log(ŷ) + (1-y)·log(1-ŷ)]，预测越准确，损失越小。'
        },
        {
            'type': '代码题',
            'question': '用 sklearn 实现逻辑回归并获取预测概率',
            'answer': '```python\nfrom sklearn.linear_model import LogisticRegression\nmodel = LogisticRegression()\nmodel.fit(X_train, y_train)\nproba = model.predict_proba(X_test)  # 返回每个类别的概率\n```'
        },
        {
            'type': '场景题',
            'question': '如何处理类别不平衡问题？',
            'answer': '1) 调整类别权重（class_weight="balanced"）；2) 重采样（过采样少数类/欠采样多数类）；3) 使用SMOTE；4) 调整分类阈值；5) 选择合适的评估指标（F1、AUC而非准确率）。'
        }
    ],

    'decision-tree': [
        {
            'type': '概念题',
            'question': '决策树如何选择最佳分裂点？',
            'answer': '通过计算不纯度的减少量。分类树用信息增益（基于熵）或基尼不纯度；回归树用MSE减少量。选择使不纯度下降最多的特征和分裂点。'
        },
        {
            'type': '对比题',
            'question': '熵 和基尼不纯度 的区别？',
            'answer': '熵考虑所有类别的概率分布，计算稍复杂；基尼不纯度计算更简单，两者效果相近。sklearn默认使用基尼。'
        },
        {
            'type': '场景题',
            'question': '决策树容易过拟合，如何解决？',
            'answer': '1) 限制树深度（max_depth）；2) 限制叶子节点最小样本数；3) 剪枝（预剪枝/后剪枝）；4) 使用随机森林等集成方法。'
        },
        {
            'type': '代码题',
            'question': '用 sklearn 训练决策树并可视化特征重要性',
            'answer': '```python\nfrom sklearn.tree import DecisionTreeClassifier\nmodel = DecisionTreeClassifier(max_depth=3, random_state=42)\nmodel.fit(X_train, y_train)\nprint(model.feature_importances_)\n```'
        },
        {
            'type': '概念题',
            'question': '预剪枝和后剪枝的区别？',
            'answer': '预剪枝在树生长过程中提前停止（如限制深度、最小样本数）；后剪枝让树完全生长后再修剪掉不重要的分支。'
        }
    ],

    'naive-bayes': [
        {
            'type': '概念题',
            'question': '朴素贝叶斯的核心假设是什么？为什么叫"朴素"？',
            'answer': '假设特征之间相互独立。之所以叫"朴素"，是因为这个假设在现实中很少成立，但算法实际效果往往很好。'
        },
        {
            'type': '对比题',
            'question': '高斯朴素贝叶斯、多项式朴素贝叶斯、伯努利朴素贝叶斯的区别？',
            'answer': '高斯NB假设特征服从正态分布（连续特征）；多项式NB适用于计数数据（如文本词频）；伯努利NB适用于二值特征（如词是否存在）。'
        },
        {
            'type': '代码题',
            'question': '用 sklearn 实现文本分类的朴素贝叶斯',
            'answer': '```python\nfrom sklearn.naive_bayes import MultinomialNB\nfrom sklearn.feature_extraction.text import CountVectorizer\n\nvectorizer = CountVectorizer()\nX_counts = vectorizer.fit_transform(texts)\nmodel = MultinomialNB()\nmodel.fit(X_counts, y)\n```'
        },
        {
            'type': '场景题',
            'question': '朴素贝叶斯在什么场景下表现特别好？',
            'answer': '1) 文本分类（垃圾邮件识别、新闻分类）；2) 实时预测（速度极快）；3) 小样本数据；4) 作为基线模型。'
        },
        {
            'type': '概念题',
            'question': '零概率问题如何解决？',
            'answer': '使用拉普拉斯平滑（Laplace Smoothing），在计算概率时给每个计数加一个小常数（通常是1），避免零概率。'
        }
    ],

    'svm': [
        {
            'type': '概念题',
            'question': '支持向量机（SVM）的核心思想是什么？',
            'answer': '寻找一个最优超平面，使得两类数据点之间的间隔（margin）最大化。支持向量是距离超平面最近的那些点。'
        },
        {
            'type': '概念题',
            'question': '核函数的作用是什么？',
            'answer': '核函数将数据映射到高维空间，使在原始空间中线性不可分的数据在高维空间中变得线性可分，无需显式计算高维坐标。'
        },
        {
            'type': '对比题',
            'question': '线性核、多项式核、RBF核的区别？',
            'answer': '线性核适用于线性可分数据，速度最快；多项式核可以拟合非线性边界；RBF核（高斯核）适用范围最广，是默认选择。'
        },
        {
            'type': '场景题',
            'question': 'SVM 对特征缩放敏感吗？为什么？',
            'answer': '非常敏感。SVM基于距离计算，如果特征尺度差异大，大尺度特征会主导距离计算。必须先进行标准化或归一化。'
        },
        {
            'type': '代码题',
            'question': '用 sklearn 实现 SVM 分类',
            'answer': '```python\nfrom sklearn.svm import SVC\nfrom sklearn.preprocessing import StandardScaler\n\nscaler = StandardScaler()\nX_scaled = scaler.fit_transform(X)\nmodel = SVC(kernel="rbf", C=1.0)\nmodel.fit(X_scaled, y)\n```'
        }
    ],

    'ensemble': [
        {
            'type': '概念题',
            'question': 'Bagging 和 Boosting 的核心区别？',
            'answer': 'Bagging（Bootstrap Aggregating）并行训练多个独立模型，通过平均/投票降低方差；Boosting 串行训练，每个新模型关注之前模型的错误，降低偏差。'
        },
        {
            'type': '概念题',
            'question': '随机森林为什么能降低过拟合？',
            'answer': '1) 每棵树使用不同的 bootstrap 样本；2) 每个分裂点只考虑部分特征（特征随机性）；3) 多棵树投票平均。'
        },
        {
            'type': '对比题',
            'question': '随机森林和 GBDT 的区别？',
            'answer': '随机森林是 Bagging，树并行生长，目标是降低方差；GBDT 是 Boosting，树串行生长，每棵树拟合残差，目标是降低偏差。'
        },
        {
            'type': '场景题',
            'question': 'XGBoost 相比传统 GBDT 有什么改进？',
            'answer': '1) 二阶梯度优化（牛顿法）；2) 正则化项防止过拟合；3) 并行处理特征分裂；4) 处理缺失值；5) 近似直方图算法加速。'
        },
        {
            'type': '代码题',
            'question': '用 sklearn 实现随机森林',
            'answer': '```python\nfrom sklearn.ensemble import RandomForestClassifier\nmodel = RandomForestClassifier(n_estimators=100, max_depth=10, random_state=42)\nmodel.fit(X_train, y_train)\n```'
        }
    ],

    'classification-metrics': [
        {
            'type': '概念题',
            'question': '准确率、精确率、召回率、F1的定义和区别？',
            'answer': '准确率 = (TP+TN)/总样本；精确率 = TP/(TP+FP)，预测为正中有多少是真的；召回率 = TP/(TP+FN)，正样本有多少被找出来；F1是精确率和召回率的调和平均。'
        },
        {
            'type': '场景题',
            'question': '在什么情况下精确率更重要？什么情况下召回率更重要？',
            'answer': '精确率重要：FP代价高（如垃圾邮件识别，不想误判正常邮件）；召回率重要：FN代价高（如疾病诊断，不想漏掉病人）。'
        },
        {
            'type': '概念题',
            'question': 'ROC曲线和AUC的含义？',
            'answer': 'ROC曲线显示TPR和FPR在不同阈值下的权衡；AUC是ROC下面积，衡量分类器整体性能，AUC=1完美，AUC=0.5随机猜测。'
        },
        {
            'type': '代码题',
            'question': '用 sklearn 计算分类报告和混淆矩阵',
            'answer': '```python\nfrom sklearn.metrics import classification_report, confusion_matrix\nprint(classification_report(y_test, y_pred))\nprint(confusion_matrix(y_test, y_pred))\n```'
        },
        {
            'type': '场景题',
            'question': '数据集类别不平衡时，应该用什么指标？',
            'answer': '不应该用准确率。应该用：F1-score（尤其关注少数类的F1）、AUC-ROC、精确率-召回率曲线、混淆矩阵分析。'
        }
    ],

    'clustering': [
        {
            'type': '概念题',
            'question': 'K-Means 的算法步骤？',
            'answer': '1) 随机初始化K个质心；2) 将每个样本分配到最近的质心；3) 重新计算每个簇的质心；4) 重复2-3直到质心不再变化或达到最大迭代次数。'
        },
        {
            'type': '场景题',
            'question': '如何选择 K-Means 的 K 值？',
            'answer': '1) 肘部法则（Elbow Method）：观察SSE下降速度的拐点；2) 轮廓系数（Silhouette Score）：衡量样本与同簇/异簇的相似度。'
        },
        {
            'type': '对比题',
            'question': 'K-Means 和 DBSCAN 的区别？',
            'answer': 'K-Means需要预设簇数量，假设簇是凸形的，对噪声敏感；DBSCAN基于密度，不需要预设簇数，可以发现任意形状的簇，能识别噪声点。'
        },
        {
            'type': '代码题',
            'question': '用 sklearn 实现 K-Means 并评估',
            'answer': '```python\nfrom sklearn.cluster import KMeans\nfrom sklearn.metrics import silhouette_score\n\nmodel = KMeans(n_clusters=3, random_state=42)\nlabels = model.fit_predict(X)\nscore = silhouette_score(X, labels)\n```'
        },
        {
            'type': '概念题',
            'question': '层次聚类和 K-Means 的区别？',
            'answer': '层次聚类构建树状结构（dendrogram），不需要预设簇数，可以可视化层次关系；K-Means更快但需要预设K值，适合大数据集。'
        }
    ],

    'pca': [
        {
            'type': '概念题',
            'question': '主成分分析（PCA）的目的是什么？',
            'answer': '降维：将高维数据投影到低维空间，同时尽可能保留数据的信息（方差）。用于可视化、加速训练、降噪。'
        },
        {
            'type': '概念题',
            'question': 'PCA 找到的主成分有什么性质？',
            'answer': '1) 主成分之间相互正交（不相关）；2) 第一主成分方向数据方差最大；3) 每个主成分都是原始特征的线性组合。'
        },
        {
            'type': '场景题',
            'question': '如何选择保留多少个主成分？',
            'answer': '1) 解释方差比例（如保留95%的方差）；2) 肘部法则观察特征值下降；3) 交叉验证看下游任务效果。'
        },
        {
            'type': '代码题',
            'question': '用 sklearn 实现 PCA',
            'answer': '```python\nfrom sklearn.decomposition import PCA\n\npca = PCA(n_components=0.95)  # 保留95%方差\nX_pca = pca.fit_transform(X)\nprint(f"原始维度: {X.shape[1]}, 降维后: {X_pca.shape[1]}")\n```'
        },
        {
            'type': '场景题',
            'question': 'PCA 之前为什么要标准化？',
            'answer': 'PCA 基于方差最大化，如果特征尺度不同，大方差的特征会主导主成分。标准化确保每个特征对等的贡献。'
        }
    ],

    'neural-networks': [
        {
            'type': '概念题',
            'question': '神经网络中的激活函数有什么作用？',
            'answer': '引入非线性，使神经网络能够拟合复杂的非线性关系。没有激活函数，多层网络等价于单层线性模型。'
        },
        {
            'type': '对比题',
            'question': 'Sigmoid、Tanh、ReLU 的区别？',
            'answer': 'Sigmoid 输出(0,1)，有梯度消失问题；Tanh 输出(-1,1)，零中心，但也有梯度消失；ReLU 计算简单，缓解梯度消失，但有神经元死亡问题。'
        },
        {
            'type': '概念题',
            'question': '什么是梯度消失？什么情况下会出现？',
            'answer': '反向传播时，梯度逐层衰减变得越来越小，导致浅层参数几乎不更新。使用 Sigmoid/Tanh 且网络较深时容易出现。'
        },
        {
            'type': '代码题',
            'question': '用 PyTorch 定义一个简单的全连接网络',
            'answer': '```python\nimport torch.nn as nn\n\nclass Net(nn.Module):\n    def __init__(self):\n        super().__init__()\n        self.fc = nn.Sequential(\n            nn.Linear(784, 256),\n            nn.ReLU(),\n            nn.Linear(256, 10)\n        )\n    def forward(self, x):\n        return self.fc(x)\n```'
        },
        {
            'type': '场景题',
            'question': '如何解决过拟合？',
            'answer': '1) Dropout；2) L1/L2 正则化（weight decay）；3) 数据增强；4) 早停；5) 减少网络规模；6) Batch Normalization。'
        }
    ],

    'optimization': [
        {
            'type': '概念题',
            'question': 'SGD（随机梯度下降）和 GD（梯度下降）的区别？',
            'answer': 'GD 使用全部数据计算梯度；SGD 使用单个样本计算梯度，更快但噪声大；Mini-batch SGD 使用一小批数据，兼顾效率和稳定性。'
        },
        {
            'type': '概念题',
            'question': '动量（Momentum）的作用是什么？',
            'answer': '积累之前的梯度方向，加速收敛，减少震荡。像物理中的惯性，帮助越过局部最优。'
        },
        {
            'type': '对比题',
            'question': 'Adam 和 SGD 的区别？',
            'answer': 'Adam 结合了动量和自适应学习率（每个参数有不同的学习率），收敛快、调参简单；SGD 泛化能力可能更好，但需要仔细调学习率。'
        },
        {
            'type': '代码题',
            'question': 'PyTorch 中使用 Adam 优化器',
            'answer': '```python\nimport torch.optim as optim\n\nmodel = Net()\noptimizer = optim.Adam(model.parameters(), lr=0.001, weight_decay=1e-5)\n# 训练循环中\noptimizer.zero_grad()\nloss.backward()\noptimizer.step()\n```'
        },
        {
            'type': '概念题',
            'question': '学习率衰减的目的是什么？',
            'answer': '训练初期用较大学习率快速收敛；后期减小学习率，让模型在最优解附近精细调整，避免震荡。常见方法：StepLR、CosineAnnealing、ReduceLROnPlateau。'
        }
    ],

    'cnn': [
        {
            'type': '概念题',
            'question': '卷积神经网络的核心组件有哪些？',
            'answer': '1) 卷积层：提取局部特征；2) 池化层：降维、平移不变性；3) 激活函数：非线性；4) 全连接层：分类。'
        },
        {
            'type': '概念题',
            'question': '感受野是什么？',
            'answer': '网络中某一层输出特征图上某个位置，对应输入图像上的区域大小。越深的层感受野越大，能看到更全局的信息。'
        },
        {
            'type': '对比题',
            'question': 'Same padding 和 Valid padding 的区别？',
            'answer': 'Same padding 填充使输出尺寸等于输入尺寸（宽/高）；Valid padding 不填充，输出尺寸变小。'
        },
        {
            'type': '代码题',
            'question': '用 PyTorch 实现一个简单的 CNN',
            'answer': '```python\nimport torch.nn as nn\n\nclass CNN(nn.Module):\n    def __init__(self):\n        super().__init__()\n        self.conv = nn.Sequential(\n            nn.Conv2d(1, 32, 3, padding=1),\n            nn.ReLU(),\n            nn.MaxPool2d(2),\n            nn.Conv2d(32, 64, 3, padding=1),\n            nn.ReLU(),\n            nn.MaxPool2d(2)\n        )\n        self.fc = nn.Linear(64*7*7, 10)\n    def forward(self, x):\n        x = self.conv(x)\n        return self.fc(x.view(x.size(0), -1))\n```'
        },
        {
            'type': '场景题',
            'question': '什么是 Batch Normalization？它有什么作用？',
            'answer': 'BN 对每个 batch 的特征进行标准化（零均值、单位方差），然后通过可学习参数恢复表示能力。作用：加速训练、允许更大学习率、缓解梯度消失、有轻微正则化效果。'
        }
    ],

    'attention': [
        {
            'type': '概念题',
            'question': 'Attention 机制的核心思想是什么？',
            'answer': '让模型在处理序列时，能够动态地关注（赋予权重）输入的不同部分，而不是固定地处理全部信息。Query-Key-Value 模式：用 Q 匹配 K 得到注意力权重，加权求和 V。'
        },
        {
            'type': '概念题',
            'question': 'Self-Attention 和 Attention 的区别？',
            'answer': 'Self-Attention 的 Q、K、V 都来自同一个输入序列，让序列内的每个位置都能注意到其他位置；普通 Attention 的 Q 和 (K,V) 来自不同序列（如编码器-解码器）。'
        },
        {
            'type': '概念题',
            'question': 'Multi-Head Attention 的作用是什么？',
            'answer': '使用多组 Q、K、V 投影，让模型能够从不同的"表示子空间"捕捉信息。每个头学习不同的注意力模式，最后拼接起来。'
        },
        {
            'type': '代码题',
            'question': 'PyTorch 中实现 Self-Attention 的核心计算',
            'answer': '```python\nimport torch\nimport torch.nn as nn\n\nclass SelfAttention(nn.Module):\n    def __init__(self, embed_dim, num_heads):\n        super().__init__()\n        self.multihead_attn = nn.MultiheadAttention(embed_dim, num_heads)\n    def forward(self, x):\n        # x: (seq_len, batch, embed_dim)\n        attn_output, _ = self.multihead_attn(x, x, x)\n        return attn_output\n```'
        },
        {
            'type': '场景题',
            'question': '为什么 Transformer 比 RNN/LSTM 更好？',
            'answer': '1) 并行计算：不依赖序列顺序，训练更快；2) 长距离依赖：Attention 直接连接任意位置，不受序列长度限制；3) 可解释性：注意力权重可视化。'
        }
    ],

    'transformer': [
        {
            'type': '概念题',
            'question': 'Transformer 的核心结构是什么？',
            'answer': 'Encoder-Decoder 架构。Encoder：输入嵌入 + 位置编码 + 多层 Self-Attention + FFN；Decoder：类似结构但多了 Encoder-Decoder Attention，输出层带 Mask 防止偷看未来。'
        },
        {
            'type': '概念题',
            'question': '位置编码（Positional Encoding）的作用？',
            'answer': 'Attention 本身不感知顺序信息，位置编码将位置信息注入输入。常用正弦/余弦编码（固定）或可学习嵌入。'
        },
        {
            'type': '场景题',
            'question': '什么是 Causal Attention / Masked Self-Attention？',
            'answer': '在 Decoder 的 Self-Attention 中，Mask 掉未来位置的信息，确保预测当前位置时只能看到之前的信息。用于自回归生成。'
        },
        {
            'type': '对比题',
            'question': 'Encoder-only、Decoder-only、Encoder-Decoder 架构的区别和应用？',
            'answer': 'Encoder-only（如 BERT）：双向理解，适合分类、标注；Decoder-only（如 GPT）：单向生成，适合文本生成；Encoder-Decoder（如 T5）：理解+生成，适合翻译、摘要。'
        },
        {
            'type': '代码题',
            'question': '使用 Hugging Face Transformers 加载模型',
            'answer': '```python\nfrom transformers import AutoTokenizer, AutoModel\n\ntokenizer = AutoTokenizer.from_pretrained("bert-base-uncased")\nmodel = AutoModel.from_pretrained("bert-base-uncased")\ninputs = tokenizer("Hello world", return_tensors="pt")\noutputs = model(**inputs)\n```'
        }
    ],

    'recommendation': [
        {
            'type': '概念题',
            'question': '协同过滤（Collaborative Filtering）的核心思想？',
            'answer': '根据用户-物品交互矩阵，利用相似用户或相似物品的偏好进行推荐。不需要物品内容信息，只需要交互历史。'
        },
        {
            'type': '对比题',
            'question': 'User-based CF 和 Item-based CF 的区别？',
            'answer': 'User-based 找相似用户，推荐他们喜欢的物品；Item-based 找相似物品，推荐与用户历史物品相似的其他物品。Item-based 通常更稳定（用户兴趣变化快，物品相似度相对固定）。'
        },
        {
            'type': '概念题',
            'question':'矩阵分解在推荐系统中的作用？',
            'answer': '将用户-物品矩阵分解为两个低秩矩阵（用户矩阵和物品矩阵）的乘积，得到用户和物品的隐向量表示，预测未交互的评分。'
        },
        {
            'type': '场景题',
            'question': '推荐系统中的冷启动问题怎么解决？',
            'answer': '1) 新用户：用内容信息（注册问卷）或推荐热门物品；2) 新物品：利用内容特征，用内容-based 方法；3) 使用混合模型；4) 探索与利用策略。'
        },
        {
            'type': '代码题',
            'question': '用 surprise 库实现协同过滤',
            'answer': '```python\nfrom surprise import SVD, Dataset, Reader\n\nreader = Reader(rating_scale=(1, 5))\ndata = Dataset.load_from_df(df[[\"user\", \"item\", \"rating\"]], reader)\ntrainset = data.build_full_trainset()\nmodel = SVD(n_factors=100)\nmodel.fit(trainset)\npred = model.predict(user_id, item_id)\n```'
        }
    ],

    'bert': [
        {
            'type': '概念题',
            'question': 'BERT 的预训练任务是什么？',
            'answer': '1) Masked Language Model（MLM）：随机mask掉15%的token，让模型预测；2) Next Sentence Prediction（NSP）：判断两句话是否连续。'
        },
        {
            'type': '对比题',
            'question': 'BERT Base 和 BERT Large 的区别？',
            'answer': 'Base：12层、768隐藏维度、12个注意力头（110M参数）；Large：24层、1024隐藏维度、16个注意力头（340M参数）。'
        },
        {
            'type': '场景题',
            'question': '如何用 BERT 做文本分类？',
            'answer': '1) 取 [CLS] token 的输出作为句子表示；2）接一个分类层（线性层）；3）Fine-tune 时可以使用较小学习率。'
        },
        {
            'type': '代码题',
            'question': '用 Hugging Face Transformers 做 BERT 分类',
            'answer': '```python\nfrom transformers import BertForSequenceClassification, BertTokenizer\n\nmodel = BertForSequenceClassification.from_pretrained("bert-base-uncased")\ntokenizer = BertTokenizer.from_pretrained("bert-base-uncased")\ninputs = tokenizer("Text here", return_tensors="pt")\noutputs = model(**inputs)\nlogits = outputs.logits\n```'
        },
        {
            'type': '概念题',
            'question': 'Word2Vec/GloVe 和 BERT 的区别？',
            'answer': 'Word2Vec/GloVe 是静态词嵌入，每个词固定表示；BERT 是上下文相关的动态嵌入，同一个词在不同上下文有不同表示。'
        }
    ],

    'gpt': [
        {
            'type': '概念题',
            'question': 'GPT 系列模型的核心特点？',
            'answer': 'Decoder-only 架构，使用自回归生成（预测下一个 token），在大规模文本上预训练，通过 Prompt 进行少样本/零样本学习。'
        },
        {
            'type': '对比题',
            'question': 'GPT-3、GPT-4 的主要改进？',
            'answer': 'GPT-3：175B参数，展示few-shot能力；GPT-4：更大规模、多模态（图文）、更长的上下文窗口、更好的推理和安全性。'
        },
        {
            'type': '概念题',
            'question': '什么是 Temperature 参数？',
            'answer': '控制生成随机性的参数。Temperature 高 → 输出更随机/有创意；Temperature 低 → 输出更确定/保守。实际是对 logits 除以 T 后再 softmax。'
        },
        {
            'type': '场景题',
            'question': 'Prompt Engineering 的基本原则？',
            'answer': '1) 清晰具体的指令；2) 提供示例（few-shot）；3) 角色设定；4) 思维链（CoT）引导；5) 输出格式约束。'
        },
        {
            'type': '代码题',
            'question': '使用 OpenAI API 调用 GPT',
            'answer': '```python\nfrom openai import OpenAI\n\nclient = OpenAI(api_key="your-key")\nresponse = client.chat.completions.create(\n    model="gpt-4",\n    messages=[{"role": "user", "content": "Hello!"}]\n)\nprint(response.choices[0].message.content)\n```'
        }
    ],

    'rag': [
        {
            'type': '概念题',
            'question': 'RAG（Retrieval-Augmented Generation）的核心思想？',
            'answer': '在生成回答前，先从知识库中检索相关文档，然后基于检索到的内容生成回答。结合了检索的准确性和生成的流畅性。'
        },
        {
            'type': '概念题',
            'question': 'RAG 的典型流程是什么？',
            'answer': '1) 文档切分；2) 用嵌入模型向量化；3) 存入向量数据库；4) 查询时将问题向量化；5) 检索最相关的文档；6) 将文档和问题一起输入LLM生成回答。'
        },
        {
            'type': '场景题',
            'question': '如何评估 RAG 系统的效果？',
            'answer': '检索质量：召回率、准确率、MRR、NDCG；生成质量：忠实度（是否基于检索内容）、相关性、流畅性；端到端：用户满意度、答案准确率。'
        },
        {
            'type': '代码题',
            'question': '用 LangChain 实现简单 RAG',
            'answer': '```python\nfrom langchain.embeddings import OpenAIEmbeddings\nfrom langchain.vectorstores import Chroma\nfrom langchain.llms import OpenAI\nfrom langchain.chains import RetrievalQA\n\nembeddings = OpenAIEmbeddings()\nvectorstore = Chroma.from_documents(docs, embeddings)\nqa = RetrievalQA.from_chain_type(llm=OpenAI(), retriever=vectorstore.as_retriever())\nanswer = qa.run("你的问题")\n```'
        },
        {
            'type': '场景题',
            'question': 'RAG 什么时候会失效？如何改进？',
            'answer': '失效：知识库中没有相关信息、检索到错误文档、LLM忽略检索内容。改进：混合检索（关键词+向量）、重排序（Rerank）、查询改写、增加检索文档数、提示词优化。'
        }
    ]
}


# ========== 格式化输出函数 ==========

def format_today_plan(plan: Dict) -> str:
    '''格式化今日计划输出'''
    separator = '═' * 50

    lines = [
        separator,
        f'📅 第{plan['week']}周 · {plan['day_name']} · Phase {plan['phase']}',
        f'{plan['phase_name']}',
        separator,
        ''
    ]

    item = plan.get('schedule_item')
    if item:
        lines.extend([
            '🌅 上午·理论 (60-90min)',
            '━━━━━━━━━━━━━━━━━━━━━━',
            item.get('morning_theory', '暂无内容'),
            ''
        ])

        # 推荐视频
        resources = plan.get('bilibili_resources', [])
        if resources:
            lines.extend([
                '🔗 推荐视频：'
            ])
            for r in resources[:3]:  # 最多显示3个
                lines.append(f'  - {r['name']} → 搜索: {r['keyword']}')
            lines.append('')

        lines.extend([
            '🌆 下午·实践 (90-120min)',
            '━━━━━━━━━━━━━━━━━━━━━━',
            item.get('afternoon_practice', '暂无内容'),
            ''
        ])

        deliverables = item.get('deliverables')
        if deliverables:
            lines.extend([
                '📦 今日交付',
                '━━━━━━━━━━━━━━━━━━━━━━',
                deliverables,
                ''
            ])

        hours = item.get('cumulative_hours')
        if hours:
            lines.extend([
                f'⏱️ 累计时长: {hours}h',
                ''
            ])

    # 今日复习卡片
    due_reviews = plan.get('due_reviews', [])
    if due_reviews:
        lines.extend([
            '📖 今日复习 (间隔重复)',
            '━━━━━━━━━━━━━━━━━━━━━━',
            f'共 {len(due_reviews)} 个概念需要复习：',
            ''
        ])
        for i, card in enumerate(due_reviews, 1):
            overdue = card.get('overdue_days', 0)
            marker = f' ⚠️过期{overdue}天' if overdue > 0 else ''
            lines.append(
                f'  {i}. {card["concept"]} '
                f'(来自W{card["source_week"]}D{card["source_day"]})'
                f'{marker}'
            )
        lines.extend([
            '',
            '💡 输入 "复习" 开始复习流程，逐个评分 0-5',
            ''
        ])

    # 学习提示
    tips = _get_learning_tips(plan['week'], plan['phase'])
    if tips:
        lines.extend([
            '💡 学习提示',
            '━━━━━━━━━━━━━━━━━━━━━━',
            tips,
            ''
        ])

    lines.append(separator)

    return '\n'.join(lines)


def _get_learning_tips(week: int, phase: int) -> str:
    '''获取个性化学习提示'''
    tips_map = {
        1: '第1周重点：建立数学直觉，不要被公式吓到！用编程思维理解矩阵运算。',
        3: '第3周有Titanic项目——这是你的第一个EDA实战，记得查看项目模板。',
        4: '线性回归是ML的基础，理解梯度下降至关重要。',
        10: '客户流失预测项目：注意处理类别不平衡问题。',
        13: '⭐ NumPy手写神经网络周！这是理解DL原理的关键一周。',
        20: '⭐ MNIST CNN ≥99% 挑战：调参的艺术，试试不同架构。',
        22: '⭐ miniGPT周：from scratch 的深度理解，准备好迎接Transformer的魅力。',
    }

    if week in tips_map:
        return tips_map[week]

    phase_tips = {
        0: 'Phase 0 是打基础的关键期，NumPy/Pandas 熟练度直接影响后续效率。',
        1: 'Phase 1 覆盖经典ML算法，每个都值得亲手实现一遍。',
        2: 'Phase 2 进入DL世界，PyTorch将成为你的主要工具。',
        3: 'Phase 3 是Transformer时代，理解Attention机制是核心。',
        4: 'Phase 4 实战LLM应用，RAG和微调是两大主线。',
        5: 'Phase 5 冲刺阶段，专注毕业项目和面试准备。'
    }

    return phase_tips.get(phase, '保持节奏，积跬步以至千里。')


def format_status(status: Dict) -> str:
    """格式化状态仪表盘输出"""
    progress_bar = "█" * int(status["progress"] / 5) + "░" * (20 - int(status["progress"] / 5))

    lines = [
        "╔══════════════════════════════════════╗",
        "║  ML/DL 学习进度仪表盘               ║",
        "╠══════════════════════════════════════╣",
        f"║  📅 当前: 第{status['current_week']}周 · {status['day_name']}               ║",
        f"║  🏷️  阶段: Phase {status['phase']} — {status['phase_name'][:20]:20s} ║",
        f"║  📊 总进度: {progress_bar} {status['progress']:.1f}%          ║",
        f"║  🔥 连续学习: {status['streak']}天                    ║",
        f"║  📁 已完成项目: {status['completed_projects']}/{status['total_projects']:2d}                 ║",
        f"║  🧪 已完成测验: {status['quiz_count']}次                    ║",
        f"║  📝 已写博客: {status['blogs']}篇                    ║",
        f"║  ⚠️  待补天数: {status['pending_makeup']}天                    ║" if status['pending_makeup'] > 0 else "║  ✅ 无待补内容                       ║",
        "╚══════════════════════════════════════╝"
    ]

    return "\n".join(lines)


def format_week_overview(overview: Dict) -> str:
    '''格式化本周概览输出'''
    lines = [
        f'═══════════════════════════════════════',
        f'📅 第{overview['week']}周概览',
        f'═══════════════════════════════════════',
        ''
    ]

    status_symbols = {
        'pending': '⬜',
        'done': '✅',
        'skipped': '⏭️ '
    }

    for day_info in overview['days']:
        symbol = status_symbols.get(day_info['status'], '⬜')
        lines.append(f'{symbol} {day_info['day_name']}')

        if day_info['schedule']:
            theory = day_info['schedule'].get('morning_theory', '')
            if theory:
                # 简化显示，只取前30个字符
                theory_short = theory[:30] + '...' if len(theory) > 30 else theory
                lines.append(f'   {theory_short}')

        lines.append('')

    lines.append('═══════════════════════════════════════')

    return '\n'.join(lines)


def format_quiz(quiz_data: Dict) -> str:
    '''格式化测验输出'''
    lines = [
        f'═══════════════════════════════════════',
        f'📝 知识检验 — {quiz_data['topic']}',
        f'═══════════════════════════════════════',
        f'',
        f'共 {quiz_data['count']} 道题，建议先思考再看答案',
        f''
    ]

    for i, q in enumerate(quiz_data['questions'], 1):
        lines.extend([
            f'【{i}】{q['question']}',
            f'   类型: {q['type']}',
            ''
        ])

        if q.get('options'):
            for opt in q['options']:
                lines.append(f'   {opt}')
            lines.append('')

    lines.extend([
        '═══════════════════════════════════════',
        '',
        '📋 答案',
        '═══════════════════════════════════════',
        ''
    ])

    for i, q in enumerate(quiz_data['questions'], 1):
        lines.extend([
            f'【{i}】{q['answer']}',
            ''
        ])

    lines.extend([
        '═══════════════════════════════════════',
        '',
        '💡 提示: 完成后记录成绩到进度中',
        ''
    ])

    return '\n'.join(lines)


def format_review(review_data: Dict) -> str:
    '''格式化周回顾输出'''
    lines = [
        f'═══════════════════════════════════════',
        f'📊 第{review_data['week']}周回顾',
        f'═══════════════════════════════════════',
        '',
        f'📈 完成情况',
        f'   ✅ 已完成: {review_data['completed']}/6 天',
        f'   ⏭️  跳过: {review_data['skipped']} 天',
        f'   ⬜ 待完成: {review_data['pending']} 天',
        f'   📊 完成率: {review_data['completion_rate']:.1f}%',
        ''
    ]

    if review_data['concepts']:
        lines.extend([
            f'📚 本周核心概念',
            ''
        ])
        for concept in review_data['concepts'][:8]:
            lines.append(f'   • {concept}')
        lines.append('')

    if review_data['quiz_questions']:
        lines.extend([
            f'📝 自测题 (建议完成后记录成绩)',
            ''
        ])
        for i, q in enumerate(review_data['quiz_questions'][:5], 1):
            lines.append(f'   {i}. {q['question']}')
        lines.append('')

    if review_data['weak_points']:
        lines.extend([
            f'⚠️  待补强内容',
            ''
        ])
        for point in review_data['weak_points']:
            lines.append(f'   [ ] {point}')
        lines.append('')

    lines.extend([
        '═══════════════════════════════════════',
        ''
    ])

    return '\n'.join(lines)


# ========== 间隔复习格式化 ==========

def format_due_reviews(cards: List[Dict]) -> str:
    '''格式化今日复习卡片列表'''
    if not cards:
        return '✅ 今日没有需要复习的概念！'

    lines = [
        '═══════════════════════════════════════',
        f'📖 今日复习卡片 — 共 {len(cards)} 个',
        '═══════════════════════════════════════',
        ''
    ]

    for i, card in enumerate(cards, 1):
        overdue = card.get('overdue_days', 0)
        marker = f' ⚠️过期{overdue}天' if overdue > 0 else ' 📅今天到期'
        ctx = card.get('source_context', '')
        ctx_line = f'     来源: {ctx}' if ctx else ''

        lines.append(f'  {i}. 【{card["concept"]}】{marker}')
        lines.append(f'     W{card["source_week"]}D{card["source_day"]} | '
                     f'已复习{card["review_count"]}次 | '
                     f'间隔{card["interval"]}天')
        if ctx_line:
            lines.append(ctx_line)
        lines.append('')

    lines.extend([
        '━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━',
        '复习方式: Claude 逐个提问，根据回答自动评分',
        '═══════════════════════════════════════',
    ])

    return '\n'.join(lines)


def format_review_result(result: Dict) -> str:
    '''格式化单次复习评分结果'''
    if 'error' in result:
        return f'❌ {result["error"]}'

    lines = [
        f'📝 {result["concept"]} — {result["status"]}',
        f'   评分: {result["quality"]}/5',
        f'   间隔: {result["old_interval"]}天 → {result["new_interval"]}天',
        f'   EF: {result["old_ef"]} → {result["new_ef"]}',
        f'   下次复习: {result["next_review"]}',
        f'   累计复习: {result["review_count"]}次',
    ]

    return '\n'.join(lines)


def format_review_stats(stats: Dict) -> str:
    '''格式化复习统计面板'''
    lines = [
        '═══════════════════════════════════════',
        '📊 间隔复习统计',
        '═══════════════════════════════════════',
        '',
        f'  📚 总卡片数: {stats["total_cards"]}',
        f'  📅 今日到期: {stats["due_today"]}',
        f'  ⚠️  已过期: {stats["overdue"]}',
        '',
        f'  🌱 新卡/年轻: {stats["young"]}',
        f'  🌳 成熟(≥21天): {stats["mature"]}',
        f'  📭 从未复习: {stats["never_reviewed"]}',
        '',
        f'  🔄 总复习次数: {stats["total_reviews"]}',
        f'  📈 平均评分: {stats["average_quality"]}/5',
        f'  📐 平均EF: {stats["average_ef"]}',
        '',
        '═══════════════════════════════════════',
    ]

    return '\n'.join(lines)


def format_learning_analytics(analytics: Dict) -> str:
    '''格式化学习分析报告'''
    if analytics['total_concepts'] == 0:
        return '📊 暂无学习数据，完成几天学习后再查看分析'

    dist = analytics['mastery_distribution']
    lines = [
        '═══════════════════════════════════════',
        '📊 学习分析报告',
        '═══════════════════════════════════════',
        '',
        f'📚 概念总数: {analytics["total_concepts"]}',
        '',
        '🎯 掌握度分布:',
        f'  😰 挣扎中 (EF<2.0): {dist["struggling"]} ({dist["struggling"]/analytics["total_concepts"]*100:.0f}%)',
        f'  📖 学习中 (EF 2.0-2.5): {dist["learning"]} ({dist["learning"]/analytics["total_concepts"]*100:.0f}%)',
        f'  ✅ 已掌握 (EF≥2.5): {dist["mastered"]} ({dist["mastered"]/analytics["total_concepts"]*100:.0f}%)',
        '',
        '📈 学习指标:',
        f'  平均复习间隔: {analytics["average_interval"]} 天',
        f'  平均 EF: {analytics["average_ef"]}',
        f'  记忆保持率估算: {analytics["retention_estimate"]}%',
        '',
        '📅 未来7天复习量预测:',
    ]

    for item in analytics['review_forecast']:
        date_obj = datetime.strptime(item['date'], '%Y-%m-%d')
        day_name = ['周一', '周二', '周三', '周四', '周五', '周六', '周日'][date_obj.weekday()]
        bar = '█' * min(item['count'], 20)
        lines.append(f'  {item["date"]} ({day_name}): {bar} {item["count"]}')

    lines.extend([
        '',
        '💡 学习建议:',
    ])

    # 根据数据给出建议
    if dist['struggling'] > analytics['total_concepts'] * 0.3:
        lines.append('  ⚠️  挣扎中的概念较多，建议降低学习速度，巩固基础')
    if analytics['retention_estimate'] < 75:
        lines.append('  ⚠️  记忆保持率偏低，建议增加复习频率')
    if max(item['count'] for item in analytics['review_forecast']) > 30:
        lines.append('  ⚠️  未来某天复习量过大，建议提前分散复习')
    if not any([dist['struggling'] > analytics['total_concepts'] * 0.3,
                analytics['retention_estimate'] < 75,
                max(item['count'] for item in analytics['review_forecast']) > 30]):
        lines.append('  ✅ 学习状态良好，保持当前节奏！')

    lines.extend([
        '',
        '═══════════════════════════════════════',
    ])

    return '\n'.join(lines)


# ========== CLI 入口点 ==========

def _get_obsidian(tutor=None):
    '''懒加载 ObsidianIntegration，复用已有 tutor 避免重复创建'''
    try:
        import sys as _sys
        _sys.path.insert(0, str(Path(__file__).parent / 'tools'))
        from obsidian_integration import ObsidianIntegration
        return ObsidianIntegration(tutor=tutor)
    except ImportError:
        return None


def main():
    '''CLI入口，用于测试'''
    import sys

    tutor = MLTutor()

    if len(sys.argv) > 1:
        cmd = sys.argv[1]

        if cmd == 'today':
            plan = tutor.get_today_plan()
            print(format_today_plan(plan))
            # 自动创建今日 Obsidian 日记
            daily_path = tutor.ensure_daily_note(plan)
            if daily_path:
                print(f'📝 今日笔记: {daily_path}')

        elif cmd == 'done':
            result = tutor.mark_done()
            if 'error' in result:
                print(f'⚠️  {result["error"]}')
                print(f'   完成时间: {result["completed_at"]}')
                print(f'📊 当前进度: {result["progress"]:.1f}%')
            else:
                print(f'✅ 第{result["week"]}周第{result["day"]}天已完成！')
                print(f'📊 总进度: {result["progress"]:.1f}%')
                print(f'🔥 连续学习: {result["streak"]}天')
                new_cards = result.get('new_review_cards', [])
                if new_cards:
                    print(f'🧠 已创建 {len(new_cards)} 张复习卡片: {", ".join(new_cards)}')

                if result.get('course_completed'):
                    print(f'\n🎓🎉 恭喜！你已完成全部 50 周的学习！')
                    print(f'   总进度: 100%')
                    print(f'   这是一个了不起的成就！')
                elif result['is_saturday']:
                    if result.get('weekly_review_generated'):
                        print(f'\n🎉 一周结束！周回顾已自动生成')
                    else:
                        completion_rate = result.get('completion_rate', 0)
                        print(f'\n⚠️  一周结束，但完成率仅 {completion_rate:.0%}')
                        print(f'   建议补做后再生成周回顾（使用 "review" 命令手动生成）')

            # 打卡后自动整理日记
            daily_path = tutor.update_daily_note_on_done(result)
            if daily_path:
                print(f'📝 今日日记已更新: {daily_path}')

            # 自动更新进度仪表盘
            obsidian = _get_obsidian(tutor)
            if obsidian:
                obsidian.update_progress_dashboard()
                print('📊 进度仪表盘已更新')

        elif cmd == 'status':
            status = tutor.get_status()
            print(format_status(status))

        elif cmd == 'week':
            overview = tutor.get_week_overview()
            print(format_week_overview(overview))

        elif cmd == 'skip' and len(sys.argv) > 2:
            reason = ' '.join(sys.argv[2:])
            result = tutor.mark_skip(reason)
            print(f'⏭️  已跳过第{result['week']}周第{result['day']}天')

        elif cmd == 'quiz':
            # 可选参数: quiz <主题>
            topic = sys.argv[2] if len(sys.argv) > 2 else None
            count = int(sys.argv[3]) if len(sys.argv) > 3 else 5
            quiz_data = tutor.generate_quiz(topic, count)
            print(format_quiz(quiz_data))
            # 自动保存测验笔记
            obsidian = _get_obsidian(tutor)
            if obsidian:
                filepath = obsidian.create_quiz_note(
                    quiz_data['topic'],
                    quiz_data['questions']
                )
                print(f'📝 测验笔记已保存: {filepath}')

        elif cmd == 'review':
            # 可选参数: review <周数>
            week = int(sys.argv[2]) if len(sys.argv) > 2 else None
            review_data = tutor.generate_weekly_review(week)
            print(format_review(review_data))
            # 保存到文件
            filepath = tutor.save_weekly_review(review_data)
            print(f'📄 回顾已保存到: {filepath}')

        elif cmd == 'save-score' and len(sys.argv) >= 4:
            # save-score <主题> <得分> <总分>
            topic = sys.argv[2]
            score = float(sys.argv[3])
            total = int(sys.argv[4])
            tutor.save_quiz_score(topic, score, total)
            print(f'✅ 成绩已保存: {score}/{total} ({topic})')

        elif cmd == 'concept' and len(sys.argv) > 2:
            concept_name = sys.argv[2]
            obsidian = _get_obsidian(tutor)
            if obsidian:
                filepath = obsidian.create_concept_note(concept_name)
                print(f'✅ 概念笔记已创建: {concept_name}')
                print(f'📝 文件位置: {filepath}')
            else:
                print('⚠️  Obsidian 集成模块未安装')

        elif cmd == 'project' and len(sys.argv) > 2:
            project_id = sys.argv[2]
            obsidian = _get_obsidian(tutor)
            if obsidian:
                filepath = obsidian.create_project_note(project_id)
                print(f'✅ 项目笔记已创建: {project_id}')
                print(f'📝 文件位置: {filepath}')
            else:
                print('⚠️  Obsidian 集成模块未安装')

        elif cmd == 'projects':
            projects = tutor.tracker.get('projects', {})
            print('📁 学习项目清单:')
            print('─' * 50)
            for pid, info in projects.items():
                symbol = {'not_started': '⬜', 'in_progress': '🔄', 'done': '✅'}.get(
                    info.get('status', 'not_started'), '⬜')
                print(f'{symbol} {pid:25s} (W{info.get("week", 1):2d})')

        elif cmd == 'dashboard':
            obsidian = _get_obsidian(tutor)
            if obsidian:
                filepath = obsidian.update_progress_dashboard()
                print(f'✅ 进度仪表盘已更新: {filepath}')
            else:
                print('⚠️  Obsidian 集成模块未安装')

        elif cmd == 'review-today':
            sr = tutor.sr_manager
            if sr is None:
                print('⚠️  间隔重复模块未安装')
            else:
                due = sr.get_due_cards()
                print(format_due_reviews(due))

        elif cmd == 'review-done' and len(sys.argv) >= 4:
            sr = tutor.sr_manager
            if sr is None:
                print('⚠️  间隔重复模块未安装')
            else:
                concept = sys.argv[2]
                quality = int(sys.argv[3])
                result = sr.review_card(concept, quality)
                print(format_review_result(result))

        elif cmd == 'review-stats':
            sr = tutor.sr_manager
            if sr is None:
                print('⚠️  间隔重复模块未安装')
            else:
                stats = sr.get_review_stats()
                print(format_review_stats(stats))

        elif cmd == 'analytics':
            sr = tutor.sr_manager
            if sr is None:
                print('⚠️  间隔重复模块未安装')
            else:
                analytics = sr.get_learning_analytics()
                print(format_learning_analytics(analytics))

        else:
            print('用法: python ml_tutor.py [命令]')
            print('')
            print('📅 每日学习:')
            print('  today           查看今日学习计划（自动创建日记）')
            print('  done            标记今日完成（自动更新仪表盘）')
            print('  status          查看总进度仪表盘')
            print('  week            查看本周概览')
            print('  skip <原因>     跳过今天')
            print('')
            print('📝 知识管理:')
            print('  quiz [主题]     生成自测题（自动保存笔记）')
            print('  review [周数]   生成周回顾')
            print('  concept <名称>  创建概念笔记')
            print('  project <ID>    创建项目笔记')
            print('  projects        列出所有项目')
            print('  dashboard       更新进度仪表盘')
            print('')
            print('📖 间隔复习:')
            print('  review-today              查看今日复习卡片')
            print('  review-done <概念> <0-5>  评分复习卡片')
            print('  review-stats              查看复习统计')
            print('  analytics                 查看学习分析报告')
    else:
        # 默认显示今日计划
        plan = tutor.get_today_plan()
        print(format_today_plan(plan))


if __name__ == '__main__':
    main()
